"""Document parsers for PDF, Word, Markdown, TXT, HTML, CSV, Excel."""
from __future__ import annotations

import io
import re
from abc import ABC, abstractmethod
from pathlib import Path


class BaseParser(ABC):
    @abstractmethod
    def parse(self, content: bytes, filename: str) -> str: ...

    @staticmethod
    def for_filename(filename: str) -> BaseParser:
        ext = Path(filename).suffix.lower()
        registry = {
            ".pdf": PDFParser(),
            ".docx": DocxParser(),
            ".doc": DocxParser(),
            ".md": MarkdownParser(),
            ".txt": TextParser(),
            ".html": HTMLParser(),
            ".htm": HTMLParser(),
            ".csv": CSVParser(),
            ".xlsx": ExcelParser(),
            ".xls": ExcelParser(),
        }
        return registry.get(ext, TextParser())


class TextParser(BaseParser):
    def parse(self, content: bytes, filename: str) -> str:
        return content.decode("utf-8", errors="replace")


class MarkdownParser(BaseParser):
    def parse(self, content: bytes, filename: str) -> str:
        return content.decode("utf-8", errors="replace")


class HTMLParser(BaseParser):
    def parse(self, content: bytes, filename: str) -> str:
        text = content.decode("utf-8", errors="replace")
        text = re.sub(r"<script[^>]*>.*?</script>", "", text, flags=re.DOTALL | re.IGNORECASE)
        text = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.DOTALL | re.IGNORECASE)
        text = re.sub(r"<[^>]+>", " ", text)
        text = re.sub(r"\s+", " ", text)
        return text.strip()


class CSVParser(BaseParser):
    def parse(self, content: bytes, filename: str) -> str:
        text = content.decode("utf-8", errors="replace")
        lines = text.strip().split("\n")
        if len(lines) <= 1:
            return text
        header = lines[0]
        result = [header]
        for line in lines[1:]:
            result.append(f"{header}: {line}")
        return "\n".join(result)


class PDFParser(BaseParser):
    def parse(self, content: bytes, filename: str) -> str:
        try:
            import fitz  # PyMuPDF
        except ImportError:
            return _fallback_text(content, filename)
        doc = fitz.open(stream=content, filetype="pdf")
        parts = []
        for page in doc:
            parts.append(page.get_text())
        doc.close()
        return "\n".join(parts)


class DocxParser(BaseParser):
    def parse(self, content: bytes, filename: str) -> str:
        try:
            from docx import Document
        except ImportError:
            return _fallback_text(content, filename)
        doc = Document(io.BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


class ExcelParser(BaseParser):
    def parse(self, content: bytes, filename: str) -> str:
        try:
            import openpyxl
        except ImportError:
            return _fallback_text(content, filename)
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"[Sheet: {sheet_name}]")
            for row in ws.iter_rows(values_only=True):
                parts.append(" | ".join(str(c) if c is not None else "" for c in row))
        wb.close()
        return "\n".join(parts)


def _fallback_text(content: bytes, filename: str) -> str:
    try:
        return content.decode("utf-8", errors="replace")
    except Exception:
        return f"[Binary file: {filename}]"
